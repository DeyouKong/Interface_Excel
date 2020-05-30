# -*- coding: utf-8 -*-

# @File: operation_xls
# @Author : "Sampson"
# @Detail :


import xlrd
import xlwt
from xlutils.copy import copy
import openpyxl
from Api_Auto_Test import *

class Operation_xls(object):
    __doc__ = "处理xls格式的excel文件"

    def __init__(self):
        self.log = log

    def write_excel_xls(self, path, sheet_name, value):
        """
        新建
        :param path:
        :param sheet_name:
        :param value:
        :return:
        """
        global sheet, workbook
        self.log.info("插入excel表格的数据：%s" % value)

        index = len(value)
        self.log.info("获取插入数据的行数：%s" % index)
        try:
            self.log.info("创建新的工作簿及新建表格")
            workbook = xlwt.Workbook()
            sheet = workbook.add_sheet(sheet_name)
        except Exception as msg:
            self.log.error("新建工作簿出错啦：%s" % msg)
        try:
            for i in range(0, index):
                for j in range(0, len(value[i])):
                    sheet.write(i, j, value[i][j])
            self.log.info("xls格式表格写入数据成功！")
            workbook.save(path)
            self.log.info("保存工作薄")
        except Exception as msg:
            self.log.error("写入数据失败：%s" % msg)



    def write_excel_xls_append(self, path, value):
        """
        在xls格式的excel文件追加数据
        :param path: 文件路径或文件名称
        :param value: 追加的数据值
        :return:
        """
        index = len(value)
        self.log.info("获取需要写入数据的行数：%s" % index)
        try:
            workbook = xlrd.open_workbook(path)  # 打开工作簿
            sheets = workbook.sheet_names()  # 获取工作簿中的所有表格
            self.log.info("打开工作簿【%s】并获取工作表格【%s】" % (workbook, sheets))
            worksheet = workbook.sheet_by_name(sheets[0])  # 获取工作簿中所有表格中的的第一个表格
            rows_old = worksheet.nrows
            self.log.info("获取表格【%s】中已存在的数据的行数：%s" % (worksheet,rows_old))
            new_workbook = copy(workbook)  # 将xlrd对象拷贝转化为xlwt对象
            new_worksheet = new_workbook.get_sheet(0)  # 获取转化后工作簿中的第一个表格
            self.log.info("【追加】写入的数据：%s" % value)
            for i in range(0, index):
                for j in range(0, len(value[i])):
                    new_worksheet.write(i + rows_old, j, value[i][j])  # 追加写入数据，注意是从i+rows_old行开始写入
            new_workbook.save(path)  # 保存工作簿
            self.log.info("xls格式表格【追加】写入数据成功！")
        except Exception as msg:
            self.log.error("打开文件或写入数据错误：%s" % msg)



    def read_excel_xls(self, path):
        """
        读取xls文件数据
        :param path: 文件路径或文件名称
        :return:
        """
        try:
            self.log.info("打开文件：%s" % path)
            workbook = xlrd.open_workbook(path)  # 打开工作簿
            sheets = workbook.sheet_names()  # 获取工作簿中的所有表格
            worksheet = workbook.sheet_by_name(sheets[0])  # 获取工作簿中所有表格中的的第一个表格
        except Exception as  msg:
            self.log.error("打开文件【%s】失败：%s" % (path, msg))
        try:
            for i in range(0, worksheet.nrows):
                for j in range(0, worksheet.ncols):
                    print(worksheet.cell_value(i, j), "\t", end="")  # 逐行逐列读取数据
                print()
        except Exception as msg:
            self.log.error("读取文件【%s】失败：%s" % (path, msg))


class Operation_xlsx(object):
    __doc__ = "处理xlsx格式的excel文件"

    def __init__(self):
        self.log = log

    def write_excel_xlsx(self, path, sheet_name, value):
        """
        数据写入xlsx格式的excel文件
        :param path: 文件路径或文件名称
        :param sheet_name: 工作簿名称
        :param value: 要写入的值： list 列表
        :return:
        """
        index = len(value)
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = sheet_name
        for i in range(0, index):
            for j in range(0, len(value[i])):
                sheet.cell(row=i + 1, column=j + 1, value=str(value[i][j]))
        workbook.save(path)
        print("xlsx格式表格写入数据成功！")

    def read_excel_xlsx(self, path, sheet_name):
        """
        读取xlsx格式的excel文件
        :param path: 文件路径或文件名称
        :param sheet_name: 工作簿名称
        :return:
        """
        workbook = openpyxl.load_workbook(path)
        # sheet = wb.get_sheet_by_name(sheet_name)这种方式已经弃用，不建议使用
        sheet = workbook[sheet_name]
        for row in sheet.rows:
            for cell in row:
                print(cell.value, "\t", end="")
            print()


if __name__ == '__main__':
    book_name_xls = 'xls格式测试工作簿.xls'
    sheet_name_xls = 'xls格式测试表'

    value_title = [["姓名", "性别", "年龄", "城市", "职业"], ]

    value1 = [["张三", "男", "19", "杭州", "研发工程师"],
              ["李四", "男", "22", "北京", "医生"],
              ["王五", "女", "33", "珠海", "出租车司机"], ]

    value2 = [["Tom", "男", "21", "西安", "测试工程师"],
              ["Jones", "女", "34", "上海", "产品经理"],
              ["Cat", "女", "56", "上海", "教师"], ]

    w = Operation_xls()
    w.write_excel_xls(book_name_xls, sheet_name_xls, value_title)
    w.write_excel_xls_append(book_name_xls, value1)
    w.write_excel_xls_append(book_name_xls, value2)
    w.read_excel_xls(book_name_xls)